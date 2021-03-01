from django.shortcuts import render, redirect, reverse
from django.http import HttpResponse
from django.views import View
from .models import *
from utils.page import PageInfo
import openpyxl
from django.forms import Form
from django.forms import fields
from django.utils.decorators import method_decorator
from django.db.models import Q
import datetime
import json
from dateutil.relativedelta import relativedelta
from django.utils.timezone import make_aware






class MyForm(Form):
    username = fields.CharField(
        max_length=16,
        min_length=6,
        required=True,
        error_messages={
            'max_length': '用户名长度应小于16位',
            'min_length': '用户名长度应大于6位',
            'required': '用户名不能为空',
        })
    password = fields.CharField(
        max_length=16,
        min_length=6,
        required=True,
        error_messages={
            'max_length': '密码长度应小于16位',
            'min_length': '密码长度应大于6位',
            'required': '密码不能为空',
        })


# Create your views here.
class Login(View):
    def get(self, request):
        return render(request, 'login.html')

    def post(self, request):
        v = MyForm(request.POST)
        if v.is_valid():
            data = UserInfo.objects.filter(username=v.cleaned_data['username']).first()
            is_remember = request.POST.get('remember')
            if is_remember:
                request.session.set_expiry(604800)
            else:
                request.session.set_expiry(0)
            if v and data.password == v.cleaned_data['password']:
                request.session['userinfo'] = {'username': v.cleaned_data['username'],
                                               'password': v.cleaned_data['password']}
                return redirect(reverse('index'))
            else:
                return render(request, 'login.html', {'erro': v.errors})
        else:
            return render(request, 'login.html', {'erro': v.errors})


def login(func):
    def wrap(request, *args, **kwargs):
        if request.session.get('userinfo'):
            return func(request, *args, **kwargs)
        else:
            return redirect(reverse('login'))

    return wrap


@login
def logout(request):
    request.session.delete()
    return redirect(reverse('login'))


@login
def management(request):
    return render(request, 'management.html')


@method_decorator(login, name='dispatch')
class Update(View):
    def get(self, request):
        return render(request, 'shebei_guanli/update.html')

    def post(self, request):
        ret = {'status': True, 'message': None}
        message_erro = "处理erro"
        try:
            file = request.FILES.get('data')
            type_taizhang = request.POST.get('ty')
            if file is None:
                return HttpResponse("请选择需要上传的文件")
            excel_type = file.name.split('.')[1]
            if excel_type in ['xlsx', 'xls']:
                wb = openpyxl.load_workbook(file)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in range(4, ws.max_row + 2):
                        if (ws.cell(row=row, column=1).value != None and ws.cell(row=row, column=2).value != None):
                            date_list = []
                            for column in range(1, ws.max_column + 1):
                                if (ws.cell(row=row, column=column).value != None):
                                    date_list.append(ws.cell(row=row, column=column).value)
                            print(date_list)
                            if (date_list):
                                if type_taizhang == 'A':
                                    A_Taizhang.objects.create(
                                        device_name=date_list[0],
                                        device_model=date_list[1],
                                        device_range=date_list[2],
                                        device_precision=date_list[3],
                                        manufacturer=date_list[4],
                                        Commissioning_date=date_list[5],
                                        device_number=date_list[6],
                                        device_factory_number=date_list[7],
                                        calibration_department=date_list[8],
                                        calibration_cycle=date_list[9],
                                        calibration_time=date_list[10],
                                        expire_time=date_list[11],
                                        jieguo_type=date_list[12],
                                        device_type=date_list[13],
                                        device_user_department=date_list[14],
                                        node=date_list[15],
                                    )
                                elif type_taizhang == 'B':
                                    B_Taizhang.objects.create(
                                        device_name=date_list[0],
                                        device_model=date_list[1],
                                        device_range=date_list[2],
                                        device_precision=date_list[3],
                                        manufacturer=date_list[4],
                                        Commissioning_date=date_list[5],
                                        device_number=date_list[6],
                                        device_factory_number=date_list[7],
                                        calibration_department=date_list[8],
                                        calibration_cycle=date_list[9],
                                        calibration_time=date_list[10],
                                        expire_time=date_list[11],
                                        jieguo_type=date_list[12],
                                        device_type=date_list[13],
                                        device_user_department=date_list[14],
                                        node=date_list[15],
                                    )
                                elif type_taizhang == 'C':
                                    C_Taizhang.objects.create(
                                        device_name=date_list[0],
                                        device_model=date_list[1],
                                        device_range=date_list[2],
                                        device_precision=date_list[3],
                                        manufacturer=date_list[4],
                                        Commissioning_date=date_list[5],
                                        device_number=date_list[6],
                                        device_factory_number=date_list[7],
                                        calibration_department=date_list[8],
                                        calibration_cycle=date_list[9],
                                        calibration_time=date_list[10],
                                        expire_time=date_list[11],
                                        jieguo_type=date_list[12],
                                        device_type=date_list[13],
                                        device_user_department=date_list[14],
                                        node=date_list[15],
                                    )


        except Exception as e:
            print(e)
            ret['status'] = False
            ret['message'] = str(e)

        return HttpResponse(json.dumps(ret))


@method_decorator(login, name='dispatch')
class Equipment_parameter(View):
    def get(self, request, shebei_type, page):
        page_info = PageInfo(page, B_Taizhang.objects.all().count(), 12, '/equipment_parameter/', 11)
        if shebei_type == 'A':
            class_list = A_Taizhang.objects.all()[page_info.start():page_info.end()]
        elif shebei_type == 'B':
            class_list = B_Taizhang.objects.all()[page_info.start():page_info.end()]
        elif shebei_type == 'C':
            class_list = C_Taizhang.objects.all()[page_info.start():page_info.end()]
        return render(request, 'shebei_guanli/shebei_index.html',
                      {'dates': class_list, 'page_info': page_info, 'shebei_type': shebei_type})

    def post(self, request, shebei_type, page):
        seacher_text = request.POST.get('search_text')
        v = B_Taizhang.objects.filter(
            Q(device_name__contains=seacher_text) |
            Q(device_model__contains=seacher_text) |
            Q(device_range__contains=seacher_text) |
            Q(device_precision__contains=seacher_text) |
            Q(manufacturer__contains=seacher_text) |
            Q(Commissioning_date__contains=seacher_text) |
            Q(device_number__contains=seacher_text) |
            Q(device_factory_number__contains=seacher_text) |
            Q(calibration_department__contains=seacher_text) |
            Q(calibration_cycle__contains=seacher_text) |
            Q(calibration_time__contains=seacher_text) |
            Q(expire_time__contains=seacher_text) |
            Q(device_type__contains=seacher_text) |
            Q(device_user_department__contains=seacher_text) |
            Q(node__contains=seacher_text)
        )
        page_info = PageInfo(page, v.count(), 12, '/equipment_parameter/', 11)
        class_list = v[page_info.start():page_info.end()]
        return render(request, 'shebei_guanli/shebei_index.html',
                      {'dates': class_list, 'page_info': page_info, 'ig': 1, 'seacher_text': seacher_text})


@method_decorator(login, name='dispatch')
class Edit(View):
    def get(self, request, shebei_type, editnumber):
        if shebei_type == 'A':
            data = A_Taizhang.objects.filter(device_factory_number=editnumber).first()
        elif shebei_type == 'B':
            data = B_Taizhang.objects.filter(device_factory_number=editnumber).first()
        elif shebei_type == 'C':
            data = C_Taizhang.objects.filter(device_factory_number=editnumber).first()
        return render(request, 'shebei_guanli/edit.html', {'data': data})

    def post(self, request, shebei_type, editnumber):
        ret = {'status': True, 'message': None}
        message_erro = "处理erro"
        try:
            datas = json.loads(request.POST.get('datas'))
            device_t = request.POST.get('device_t')
            if device_t == 'B':
                if datas['device_type'] == 'B':
                    B_Taizhang.objects.filter(device_factory_number=editnumber).update(**datas)
                else:
                    B_Taizhang.objects.filter(device_factory_number=editnumber).delete()
                    A_Taizhang.objects.create(**datas)
            else:
                if datas['device_type'] == 'B':
                    A_Taizhang.objects.filter(device_factory_number=editnumber).delete()
                    B_Taizhang.objects.create(**datas)
                else:
                    A_Taizhang.objects.filter(device_factory_number=editnumber).update(**datas)

        except Exception as e:
            ret['status'] = False
            ret['message'] = str(e)

        return HttpResponse(json.dumps(ret))


@login
def index(request):
    v = request.session.get('userinfo')
    user_info = UserInfo.objects.filter(username=v['username'], password=v['password']).first()
    now_data = datetime.datetime.now().strftime('%Y-%m-%d')
    add_day = datetime.datetime.now() + datetime.timedelta(days=7)
    A_today_dates = B_Taizhang.objects.filter(expire_time=now_data)
    B_today_dates = A_Taizhang.objects.filter(expire_time=now_data)

    A_day_7_dates = A_Taizhang.objects.filter(expire_time__lt=add_day.strftime('%Y-%m-%d'))
    B_day_7_dates = B_Taizhang.objects.filter(expire_time__lt=add_day.strftime('%Y-%m-%d'))

    shebei_count = A_Taizhang.objects.count() + B_Taizhang.objects.count() + C_Taizhang.objects.count()


    def last_day_of_month(any_day):
        """
        获取获得一个月中的最后一天
        :param any_day: 任意日期
        :return: string
        """
        next_month = any_day.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
        return next_month - datetime.timedelta(days=next_month.day)

    last_day = last_day_of_month(datetime.datetime.now())

    startday = datetime.datetime(year=2021,month=2,day=1).strftime('%Y-%m-%d')

    dangyueAleisum = A_Taizhang.objects.filter(expire_time__range=(startday,last_day))
    dangyueBleisum = B_Taizhang.objects.filter(expire_time__range=(startday,last_day))
    now_month = 'weiwancheng_'+ str(datetime.datetime.now().month)
    now_wanmonth = 'wancheng_'+ str(datetime.datetime.now().month)
    v = Meiyuewanchengshu.objects.filter(now_year=str(datetime.datetime.now().year))
    if v:
        v.update(now_month=dangyueAleisum.count() + dangyueBleisum.count())
    else:
        Meiyuewanchengshu.objects.create(now_year=str(datetime.datetime.now().year))
        Meiyuewanchengshu.objects.filter(now_year=str(datetime.datetime.now().year)).update(now_month=dangyueAleisum.count() + dangyueBleisum.count())
    # str(datetime.datetime.now().year)
    # shebeimeiyuewanchengshu[datetime.datetime.now().month][0] = dangyueAleisum.count() + dangyueBleisum.count()
    # qunianwanchengsum = 0
    # for i in range(len(shebeimeiyuewanchengshu)):
    #     qunianwanchengsum += shebeimeiyuewanchengshu[datetime.datetime.now().month][1]

    return render(request, 'index.html',
                  {'A_today_dates': A_today_dates, 'B_today_dates': B_today_dates, 'A_day_7_dates': A_day_7_dates,
                   'B_day_7_dates': B_day_7_dates, 'user_info': user_info, 'shebei_count': shebei_count,
                   })


@login
def test(request):
    if request.method == 'GET':
        return render(request, 'test.html')
    else:
        r = request.POST.get('dates')
        print(r)
        return HttpResponse(r)


@method_decorator(login, name='dispatch')
class Profile(View):
    def get(self, request):
        v = request.session.get('userinfo')
        user_info = UserInfo.objects.filter(username=v['username'], password=v['password']).first()
        return render(request, 'profile.html', {'user_info': user_info})

    def post(self, request):
        ret = {'status': True, 'message': None}
        message_erro = "处理erro"
        try:
            username = request.POST.get('username')
            nickname = request.POST.get('nickname')
            email = request.POST.get('email')
            abstract = request.POST.get('abstract')
            UserInfo.objects.filter(username=username).update(nickname=nickname, email=email, abstract=abstract)
        except Exception as e:
            ret['status'] = False
            ret['message'] = str(e)

        return HttpResponse(json.dumps(ret))


@method_decorator(login, name='dispatch')
class Edit_pwd(View):
    def get(self, request):
        return render(request, 'edit_pwd.html')

    def post(self, request):
        ret = {'status': True, 'message': None}
        message_erro = "处理erro"
        try:
            old_password = request.POST.get('old_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            v = request.session.get('userinfo')
            user_info = UserInfo.objects.filter(username=v['username'], password=v['password']).first()
            if user_info.password == old_password and user_info.email == confirm_password:
                UserInfo.objects.filter(username=v['username'], password=v['password']).update(password=new_password)
            else:
                raise Exception("密码输入错误，或邮箱错误！")
        except Exception as e:
            ret['status'] = False
            ret['message'] = str(e)

        return HttpResponse(json.dumps(ret))


@login
def complete(request, shebei_type, editnumber):
    if shebei_type == 'A':
        v = A_Taizhang.objects.filter(device_factory_number=editnumber).first()
    elif shebei_type == 'B':
        v = B_Taizhang.objects.filter(device_factory_number=editnumber).first()
    now_time = datetime.datetime.now()
    calibration_time = now_time.strftime('%Y-%m-%d')
    if v.calibration_cycle == '半年':
        expire_time = now_time + relativedelta(months=+6) - relativedelta(days=+1)
        print(expire_time)
    elif v.calibration_cycle == '1年':
        expire_time = now_time + relativedelta(years=+1) - relativedelta(days=+1)
        print(expire_time)
    elif v.calibration_cycle == '2年':
        expire_time = now_time + relativedelta(years=+2) - relativedelta(days=+1)
        print(expire_time)
    if shebei_type == 'A':
        A_Taizhang.objects.filter(device_factory_number=editnumber).update(calibration_time=calibration_time,
                                                                           expire_time=expire_time)
    elif shebei_type == 'B':
        B_Taizhang.objects.filter(device_factory_number=editnumber).update(calibration_time=calibration_time,
                                                                           expire_time=expire_time)
    shebeimeiyuewanchengshu[datetime.datetime.now().month][1] += 1
    return redirect(reverse('equipment_parameter', kwargs={'shebei_type': shebei_type, 'page': 1}))


@login
def completes(request):
    ret = {'status': True, 'message': None}
    message_erro = "处理erro"
    try:
        v = request.POST.get('values').split(',')
        for dev in v:
            shebei_type, editnumber = dev.split(':')
            if shebei_type == 'A':
                v = A_Taizhang.objects.filter(device_factory_number=editnumber).first()
            elif shebei_type == 'B':
                v = B_Taizhang.objects.filter(device_factory_number=editnumber).first()
            now_time = datetime.datetime.now()
            calibration_time = now_time.strftime('%Y-%m-%d')
            if v.calibration_cycle == '半年':
                expire_time = now_time + relativedelta(months=+6) - relativedelta(days=+1)
                print(expire_time)
            elif v.calibration_cycle == '1年':
                expire_time = now_time + relativedelta(years=+1) - relativedelta(days=+1)
                print(expire_time)
            elif v.calibration_cycle == '2年':
                expire_time = now_time + relativedelta(years=+2) - relativedelta(days=+1)
                print(expire_time)
            if shebei_type == 'A':
                A_Taizhang.objects.filter(device_factory_number=editnumber).update(calibration_time=calibration_time,
                                                                                   expire_time=expire_time)
            elif shebei_type == 'B':
                B_Taizhang.objects.filter(device_factory_number=editnumber).update(calibration_time=calibration_time,
                                                                                   expire_time=expire_time)
            shebeimeiyuewanchengshu[datetime.datetime.now().month][1] += 1

    except Exception as e:
        ret['status'] = False
        ret['message'] = str(e)

    return HttpResponse(json.dumps(ret))
